"""
Simon's Magic Data Machine
---------------------------
A Streamlit dashboard that turns a Delphi (Salesforce) events-booking
export into financial and pipeline KPIs for The Brewery.

Run locally:   streamlit run app.py
Deploy:        see README.md
"""

import io
from datetime import datetime

import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# PAGE CONFIG
# --------------------------------------------------------------------------
st.set_page_config(
    page_title="Simon's Magic Data Machine",
    page_icon="🍺",
    layout="wide",
)

# --------------------------------------------------------------------------
# CONSTANTS — column mapping & business rules
# --------------------------------------------------------------------------
# Maps the raw Delphi/Salesforce export column names to clean internal names.
# If your export ever renames a column slightly, add the new name here.
COLUMN_MAP = {
    "Booking: Booking Post As": "Booking_ID",
    "Account": "Account",
    "Status": "Status",
    "Blended Revenue Total Currency": "Currency",
    "Blended Revenue Total": "Revenue",
    "Arrival": "Arrival_Date",
    "Booking: Created Date": "Created_Date",
    "Booking: Owner Alias": "Sales_Rep",
    "Date Definite": "Date_Definite",
    "Date Lost": "Date_Lost",
    "Date Tentative": "Date_Tentative",
    "Lead Source": "Lead_Source",
    "Meeting Class": "Meeting_Class",
    "Date Turned Down": "Date_Turned_Down",
    "Lost Reason": "Lost_Reason",
}

DATE_COLS = [
    "Arrival_Date", "Created_Date", "Date_Definite",
    "Date_Lost", "Date_Tentative", "Date_Turned_Down",
]

# Business rules confirmed with Simon:
#   - Won            = Definite
#   - Lost (excluded)= Lost, TurnedDown, Cancelled  (Cancelled counts as lost revenue)
#   - Open pipeline  = Tentative, Prospect (Prospect == "Provisional")
WON_STATUS = "Definite"
LOST_STATUSES = ["Lost", "TurnedDown", "Cancelled"]
OPEN_STATUSES = ["Tentative", "Prospect"]


# --------------------------------------------------------------------------
# DATA LOADING
# --------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def load_data(uploaded_file) -> pd.DataFrame:
    """
    Reads the uploaded file. Delphi/Salesforce 'xls' exports are frequently
    actually HTML tables wearing an .xls extension, so we try real Excel
    formats first, then fall back to HTML parsing.
    """
    # Try modern Excel (.xlsx)
    try:
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, engine="openpyxl")
        return df
    except Exception:
        pass

    # Try legacy Excel (.xls, true binary format)
    try:
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, engine="xlrd")
        return df
    except Exception:
        pass

    # Fall back to HTML-disguised-as-xls (common Salesforce/Delphi export)
    try:
        uploaded_file.seek(0)
        tables = pd.read_html(uploaded_file)
        if tables:
            return tables[0]
    except Exception:
        pass

    raise ValueError(
        "Could not read this file. Please make sure it's a Delphi export "
        "in .xlsx or .xls format."
    )


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Renames columns, parses dates, and derives helper fields."""
    # Rename any columns we recognise; leave others untouched
    rename_dict = {k: v for k, v in COLUMN_MAP.items() if k in df.columns}
    df = df.rename(columns=rename_dict)

    missing = [v for v in COLUMN_MAP.values() if v not in df.columns]
    if missing:
        st.warning(
            f"Heads up — these expected columns weren't found in your file "
            f"and related metrics will be skipped: {', '.join(missing)}"
        )

    # Parse dates (Delphi exports as dd/mm/yyyy)
    for col in DATE_COLS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], format="%d/%m/%Y", errors="coerce")

    # Ensure Revenue is numeric
    if "Revenue" in df.columns:
        df["Revenue"] = pd.to_numeric(df["Revenue"], errors="coerce").fillna(0)

    # Derive Financial Year from Arrival Date (FY runs 1 July - 30 June)
    if "Arrival_Date" in df.columns:
        df["Financial_Year"] = df["Arrival_Date"].apply(fy_from_date)

    # Derive a simple outcome bucket for easy filtering/charting
    if "Status" in df.columns:
        df["Outcome"] = df["Status"].apply(bucket_status)

    return df


def fy_from_date(d) -> str:
    """Converts a date into a 'FY2025/26' style label (FY = 1 Jul - 30 Jun)."""
    if pd.isna(d):
        return "Unknown"
    start_year = d.year if d.month >= 7 else d.year - 1
    return f"FY{start_year}/{str(start_year + 1)[-2:]}"


def bucket_status(status: str) -> str:
    if status == WON_STATUS:
        return "Won"
    if status in LOST_STATUSES:
        return "Lost"
    if status in OPEN_STATUSES:
        return "Open"
    return "Other"


# --------------------------------------------------------------------------
# KPI CALCULATIONS
# --------------------------------------------------------------------------
def calculate_kpis(df: pd.DataFrame) -> dict:
    k = {}

    total_leads = len(df)
    won_df = df[df["Status"] == WON_STATUS]
    lost_df = df[df["Status"].isin(LOST_STATUSES)]
    open_df = df[df["Status"].isin(OPEN_STATUSES)]

    won_revenue = won_df["Revenue"].sum()
    total_pipeline_value = df["Revenue"].sum()

    # --- Core metrics ---
    k["total_leads"] = total_leads
    k["won_revenue"] = won_revenue
    k["total_pipeline_value"] = total_pipeline_value
    k["revenue_per_lead"] = won_revenue / total_leads if total_leads else 0
    k["pipeline_conversion"] = (
        won_revenue / total_pipeline_value if total_pipeline_value else 0
    )
    k["conversion_rate"] = len(won_df) / total_leads if total_leads else 0

    # --- Additional KPIs ---
    # 1. Average Deal Size (Won only)
    k["avg_deal_size"] = won_revenue / len(won_df) if len(won_df) else 0

    # 2. Sales Cycle Length: enquiry -> signed contract, for Won deals
    cycle_df = won_df.dropna(subset=["Created_Date", "Date_Definite"])
    if len(cycle_df):
        cycle_days = (cycle_df["Date_Definite"] - cycle_df["Created_Date"]).dt.days
        k["avg_sales_cycle_days"] = cycle_days.mean()
    else:
        k["avg_sales_cycle_days"] = None

    # 3. Lost vs Turned Down split (both are "not won" but mean different things)
    lost_count = (df["Status"] == "Lost").sum()
    turned_down_count = (df["Status"] == "TurnedDown").sum()
    not_won_total = lost_count + turned_down_count
    k["lost_rate"] = lost_count / not_won_total if not_won_total else 0
    k["turned_down_rate"] = turned_down_count / not_won_total if not_won_total else 0
    k["lost_count"] = lost_count
    k["turned_down_count"] = turned_down_count

    # 4. Cancellation Rate: signed deals that later fell through
    cancelled_count = (df["Status"] == "Cancelled").sum()
    signed_total = len(won_df) + cancelled_count
    k["cancellation_rate"] = cancelled_count / signed_total if signed_total else 0
    k["cancelled_count"] = cancelled_count

    # 5. Win rate by Lead Source
    if "Lead_Source" in df.columns:
        by_source = df.groupby("Lead_Source").agg(
            Total_Leads=("Status", "count"),
            Won=("Status", lambda s: (s == WON_STATUS).sum()),
            Revenue_Won=("Revenue", lambda s: s[df.loc[s.index, "Status"] == WON_STATUS].sum()),
        )
        by_source["Win_Rate"] = by_source["Won"] / by_source["Total_Leads"]
        k["by_lead_source"] = by_source.sort_values("Revenue_Won", ascending=False)

    # 6. Sales Rep Leaderboard — Total leads handled, won, win rate, revenue, avg deal size
    if "Sales_Rep" in df.columns:
        leaderboard = df.groupby("Sales_Rep").agg(
            Total_Leads=("Status", "count"),
            Won_Deals=("Status", lambda s: (s == WON_STATUS).sum()),
        )
        revenue_by_rep = won_df.groupby("Sales_Rep")["Revenue"].sum()
        leaderboard["Revenue_Won"] = leaderboard.index.map(revenue_by_rep).fillna(0)
        leaderboard["Win_Rate"] = leaderboard["Won_Deals"] / leaderboard["Total_Leads"]
        leaderboard["Avg_Deal_Size"] = leaderboard.apply(
            lambda r: r["Revenue_Won"] / r["Won_Deals"] if r["Won_Deals"] else 0, axis=1
        )
        leaderboard = leaderboard.sort_values("Revenue_Won", ascending=False)
        leaderboard.insert(0, "Rank", range(1, len(leaderboard) + 1))
        k["leaderboard"] = leaderboard

    # 7. Revenue AND Win Rate by Meeting Class (event type)
    if "Meeting_Class" in df.columns:
        by_class = df.groupby("Meeting_Class").agg(
            Total_Leads=("Status", "count"),
            Won_Deals=("Status", lambda s: (s == WON_STATUS).sum()),
            Revenue_Won=("Revenue", lambda s: s[df.loc[s.index, "Status"] == WON_STATUS].sum()),
        )
        by_class["Win_Rate"] = by_class["Won_Deals"] / by_class["Total_Leads"]
        k["by_meeting_class"] = by_class.sort_values("Revenue_Won", ascending=False)

    # 8. Pipeline Aging — open bookings bucketed by days since enquiry
    if len(open_df):
        today = pd.Timestamp(datetime.now().date())
        aging = (today - open_df["Created_Date"]).dt.days
        bins = [-1, 30, 60, 90, 10_000]
        labels = ["0-30 days", "31-60 days", "61-90 days", "90+ days"]
        aging_bucket = pd.cut(aging, bins=bins, labels=labels)
        k["pipeline_aging"] = (
            open_df.assign(Aging_Bucket=aging_bucket)
            .groupby("Aging_Bucket", observed=True)
            .agg(Open_Deals=("Status", "count"), Open_Value=("Revenue", "sum"))
        )
    else:
        k["pipeline_aging"] = None

    # 9. Seasonality — bookings/revenue by calendar month of the event (Arrival),
    #    ordered by financial year (Jul -> Jun) so it reads naturally for The Brewery
    if "Arrival_Date" in df.columns:
        month_df = df.dropna(subset=["Arrival_Date"]).copy()
        if len(month_df):
            month_df["Month_Num"] = month_df["Arrival_Date"].dt.month
            month_df["Month_Name"] = month_df["Arrival_Date"].dt.strftime("%B")
            seasonality = month_df.groupby(["Month_Num", "Month_Name"]).agg(
                Total_Bookings=("Status", "count"),
                Won_Deals=("Status", lambda s: (s == WON_STATUS).sum()),
                Revenue_Won=(
                    "Revenue",
                    lambda s: s[month_df.loc[s.index, "Status"] == WON_STATUS].sum(),
                ),
            ).reset_index()
            # Reorder Jul(7)..Jun(6) to match the financial year instead of Jan..Dec
            seasonality["FY_Order"] = seasonality["Month_Num"].apply(
                lambda m: m - 7 if m >= 7 else m + 5
            )
            seasonality = (
                seasonality.sort_values("FY_Order")
                .drop(columns=["FY_Order", "Month_Num"])
                .set_index("Month_Name")
            )
            k["seasonality"] = seasonality
        else:
            k["seasonality"] = None
    else:
        k["seasonality"] = None

    # 10. Revenue trend over time — monthly, based on event (Arrival) month.
    #     Shows momentum/decline that a single financial-year snapshot can hide.
    if "Arrival_Date" in df.columns:
        trend_df = df.dropna(subset=["Arrival_Date"]).copy()
        if len(trend_df):
            trend_df["Year_Month"] = trend_df["Arrival_Date"].dt.to_period("M").astype(str)
            revenue_trend = trend_df.groupby("Year_Month").agg(
                Total_Bookings=("Status", "count"),
                Won_Deals=("Status", lambda s: (s == WON_STATUS).sum()),
                Revenue_Won=(
                    "Revenue",
                    lambda s: s[trend_df.loc[s.index, "Status"] == WON_STATUS].sum(),
                ),
            ).sort_index()
            k["revenue_trend"] = revenue_trend
        else:
            k["revenue_trend"] = None
    else:
        k["revenue_trend"] = None

    # 11. Account concentration risk — how much of Won revenue sits in the top 10 accounts
    if "Account" in df.columns:
        account_revenue = won_df.groupby("Account")["Revenue"].sum().sort_values(ascending=False)
        top10 = account_revenue.head(10)
        k["top10_account_share"] = top10.sum() / won_revenue if won_revenue else 0
        k["top_accounts"] = (
            top10.reset_index().rename(columns={"Revenue": "Revenue_Won"})
        )
    else:
        k["top10_account_share"] = None
        k["top_accounts"] = None

    # 12. Lost Reason analysis — covers both Lost ("LT-" prefix) and TurnedDown ("TD-" prefix)
    #     bookings, since Reason Lost is populated for both outcome types
    if "Lost_Reason" in df.columns:
        reason_df = df[df["Status"].isin(["Lost", "TurnedDown"])].dropna(subset=["Lost_Reason"])
        if len(reason_df):
            k["lost_reasons"] = (
                reason_df.groupby("Lost_Reason")
                .agg(Count=("Status", "count"), Value=("Revenue", "sum"))
                .sort_values("Count", ascending=False)
            )
        else:
            k["lost_reasons"] = None
    else:
        k["lost_reasons"] = None

    return k


# --------------------------------------------------------------------------
# EXCEL EXPORT
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="2C1B0E", end_color="2C1B0E", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
LABEL_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=11)


def _style_header_row(ws, row_num, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws, df, start_col=1):
    for i, col in enumerate(df.columns):
        col_letter = get_column_letter(start_col + i)
        max_len = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def build_excel_report(df: pd.DataFrame, kpis: dict, fy_label: str, rep_label: str) -> bytes:
    """Builds a multi-sheet Excel workbook summarising the current KPI view."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ---- Sheet 1: Summary ----
        summary_rows = [
            ("Report", "Simon's Magic Data Machine — KPI Summary"),
            ("Generated On", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("Financial Year Filter", fy_label),
            ("Sales Person Filter", rep_label),
            ("", ""),
            ("CORE METRICS", ""),
            ("Total Leads (Bookings)", kpis["total_leads"]),
            ("Total Pipeline Value (£)", round(kpis["total_pipeline_value"], 2)),
            ("Actual Revenue Won (£)", round(kpis["won_revenue"], 2)),
            ("Revenue per Lead (£)", round(kpis["revenue_per_lead"], 2)),
            ("Pipeline Value Conversion (%)", round(kpis["pipeline_conversion"] * 100, 2)),
            ("Conversion Rate (%)", round(kpis["conversion_rate"] * 100, 2)),
            ("", ""),
            ("ADDITIONAL KPIs", ""),
            ("Average Deal Size - Won (£)", round(kpis["avg_deal_size"], 2)),
            (
                "Average Sales Cycle (days)",
                round(kpis["avg_sales_cycle_days"], 1) if kpis["avg_sales_cycle_days"] is not None else "N/A",
            ),
            ("Lost to Competitor Rate (%)", round(kpis["lost_rate"] * 100, 2)),
            ("Turned Down Rate (%)", round(kpis["turned_down_rate"] * 100, 2)),
            ("Cancellation Rate (%)", round(kpis["cancellation_rate"] * 100, 2)),
            (
                "Top 10 Accounts' Share of Won Revenue (%)",
                round(kpis["top10_account_share"] * 100, 2) if kpis.get("top10_account_share") is not None else "N/A",
            ),
        ]
        summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=0)

        ws = writer.sheets["Summary"]
        ws.cell(row=1, column=1).font = TITLE_FONT
        _style_header_row(ws, 1, 2)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            row[0].font = LABEL_FONT
        _autofit_columns(ws, summary_df)

        # ---- Sheet 2: Sales Rep Leaderboard ----
        if "leaderboard" in kpis:
            lb_df = kpis["leaderboard"].reset_index()
            lb_df.to_excel(writer, sheet_name="Leaderboard", index=False)
            ws = writer.sheets["Leaderboard"]
            _style_header_row(ws, 1, len(lb_df.columns))
            _autofit_columns(ws, lb_df)

        # ---- Sheet 3: By Lead Source ----
        if "by_lead_source" in kpis:
            src_df = kpis["by_lead_source"].reset_index()
            src_df.to_excel(writer, sheet_name="By Lead Source", index=False)
            ws = writer.sheets["By Lead Source"]
            _style_header_row(ws, 1, len(src_df.columns))
            _autofit_columns(ws, src_df)

        # ---- Sheet 4: By Event Type ----
        if "by_meeting_class" in kpis:
            class_df = kpis["by_meeting_class"].reset_index()
            class_df.to_excel(writer, sheet_name="By Event Type", index=False)
            ws = writer.sheets["By Event Type"]
            _style_header_row(ws, 1, len(class_df.columns))
            _autofit_columns(ws, class_df)

        # ---- Sheet 5: Pipeline Aging ----
        if kpis.get("pipeline_aging") is not None:
            aging_df = kpis["pipeline_aging"].reset_index()
            aging_df.to_excel(writer, sheet_name="Pipeline Aging", index=False)
            ws = writer.sheets["Pipeline Aging"]
            _style_header_row(ws, 1, len(aging_df.columns))
            _autofit_columns(ws, aging_df)

        # ---- Sheet 6: Seasonality ----
        if kpis.get("seasonality") is not None:
            season_df = kpis["seasonality"].reset_index()
            season_df.to_excel(writer, sheet_name="Seasonality", index=False)
            ws = writer.sheets["Seasonality"]
            _style_header_row(ws, 1, len(season_df.columns))
            _autofit_columns(ws, season_df)

        # ---- Sheet 7: Revenue Trend ----
        if kpis.get("revenue_trend") is not None:
            trend_df = kpis["revenue_trend"].reset_index()
            trend_df.to_excel(writer, sheet_name="Revenue Trend", index=False)
            ws = writer.sheets["Revenue Trend"]
            _style_header_row(ws, 1, len(trend_df.columns))
            _autofit_columns(ws, trend_df)

        # ---- Sheet 8: Top 10 Accounts ----
        if kpis.get("top_accounts") is not None:
            acct_df = kpis["top_accounts"]
            acct_df.to_excel(writer, sheet_name="Top Accounts", index=False)
            ws = writer.sheets["Top Accounts"]
            _style_header_row(ws, 1, len(acct_df.columns))
            _autofit_columns(ws, acct_df)
            note_row = len(acct_df) + 3
            ws.cell(row=note_row, column=1, value="Top 10 accounts' share of total Won revenue:")
            ws.cell(row=note_row, column=2, value=f"{kpis['top10_account_share']*100:.1f}%")

        # ---- Sheet 9: Lost / Turned Down Reasons ----
        if kpis.get("lost_reasons") is not None:
            reasons_df = kpis["lost_reasons"].reset_index()
            reasons_df.to_excel(writer, sheet_name="Lost Reasons", index=False)
            ws = writer.sheets["Lost Reasons"]
            _style_header_row(ws, 1, len(reasons_df.columns))
            _autofit_columns(ws, reasons_df)

    return output.getvalue()


# --------------------------------------------------------------------------
# UI HELPERS
# --------------------------------------------------------------------------
def gbp(value) -> str:
    if value is None or pd.isna(value):
        return "—"
    return f"£{value:,.0f}"


def pct(value) -> str:
    if value is None or pd.isna(value):
        return "—"
    return f"{value * 100:.1f}%"


# --------------------------------------------------------------------------
# MAIN APP
# --------------------------------------------------------------------------
def main():
    st.title("🍺 Simon's Magic Data Machine")
    st.caption(
        "Upload your Delphi booking export to see The Brewery's sales "
        "pipeline and revenue KPIs — instantly."
    )

    uploaded_file = st.file_uploader(
        "Upload your Delphi export (.xlsx or .xls)", type=["xlsx", "xls"]
    )

    if uploaded_file is None:
        st.info("👆 Upload a file to get started.")
        st.stop()

    with st.spinner("Reading your file..."):
        try:
            raw_df = load_data(uploaded_file)
            df = clean_data(raw_df)
        except Exception as e:
            st.error(f"Something went wrong reading this file: {e}")
            st.stop()

    # ---------------- Sidebar filters ----------------
    st.sidebar.header("Filters")

    fy_options = ["All Time"] + sorted(
        [fy for fy in df["Financial_Year"].dropna().unique() if fy != "Unknown"],
        reverse=True,
    )
    selected_fy = st.sidebar.selectbox("Financial Year (based on Arrival date)", fy_options)

    if selected_fy != "All Time":
        df = df[df["Financial_Year"] == selected_fy]

    selected_rep = "All"
    if "Sales_Rep" in df.columns:
        reps = ["All"] + sorted(df["Sales_Rep"].dropna().unique().tolist())
        selected_rep = st.sidebar.selectbox("Sales Person", reps)
        if selected_rep != "All":
            df = df[df["Sales_Rep"] == selected_rep]

    if len(df) == 0:
        st.warning("No bookings match the selected filters.")
        st.stop()

    kpis = calculate_kpis(df)

    # ---------------- Core Metrics ----------------
    st.header("Core Metrics")
    c1, c2, c3 = st.columns(3)
    c1.metric(
        "Revenue per Lead",
        gbp(kpis["revenue_per_lead"]),
        help="Won Revenue ÷ Total Number of Bookings (leads) in the selected period.",
    )
    c2.metric(
        "Pipeline Value Conversion",
        pct(kpis["pipeline_conversion"]),
        help="Actual Revenue Won ÷ Total Pipeline Value (sum of all booking values, Won + Lost + Open).",
    )
    c3.metric(
        "Conversion Rate",
        pct(kpis["conversion_rate"]),
        help="Number of Definite (Won) bookings ÷ Total Number of Bookings (leads).",
    )

    st.caption(
        f"Based on {kpis['total_leads']:,} bookings totalling "
        f"{gbp(kpis['total_pipeline_value'])} in pipeline value, of which "
        f"{gbp(kpis['won_revenue'])} was won."
    )

    excel_bytes = build_excel_report(df, kpis, selected_fy, selected_rep)
    file_label = selected_fy.replace("/", "-") if selected_fy != "All Time" else "All-Time"
    st.download_button(
        label="📥 Download KPI Summary (Excel)",
        data=excel_bytes,
        file_name=f"Brewery_KPI_Summary_{file_label}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.divider()

    # ---------------- Additional KPIs ----------------
    st.header("Additional KPIs")
    c1, c2, c3 = st.columns(3)
    c1.metric("Average Deal Size (Won)", gbp(kpis["avg_deal_size"]))
    c2.metric(
        "Avg. Sales Cycle",
        f"{kpis['avg_sales_cycle_days']:.0f} days" if kpis["avg_sales_cycle_days"] is not None else "—",
        help="Average days from enquiry (Booking Created Date) to signed contract (Date Definite).",
    )
    c3.metric(
        "Cancellation Rate",
        pct(kpis["cancellation_rate"]),
        help="Of all signed deals (Definite + Cancelled), the % that later cancelled.",
    )

    c4, c5, c6 = st.columns(3)
    c4.metric(
        "Lost to Competitor Rate",
        pct(kpis["lost_rate"]),
        f"{kpis['lost_count']:,} bookings",
        help="Of all non-won outcomes (Lost + TurnedDown), the % where the client chose someone else.",
    )
    c5.metric(
        "Turned Down Rate",
        pct(kpis["turned_down_rate"]),
        f"{kpis['turned_down_count']:,} bookings",
        help="Of all non-won outcomes (Lost + TurnedDown), the % where The Brewery declined to host.",
    )
    c6.metric(
        "Top 10 Account Concentration",
        pct(kpis["top10_account_share"]) if kpis.get("top10_account_share") is not None else "—",
        help="Share of total Won revenue that comes from your top 10 accounts. High values mean more exposure if one client leaves.",
    )

    st.divider()

    # ---------------- Charts ----------------
    st.header("Breakdowns")

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs(
        [
            "Sales Leaderboard", "By Lead Source", "By Event Type", "Pipeline Aging",
            "Seasonality", "Revenue Trend", "Lost Reasons", "Top Accounts",
        ]
    )

    with tab1:
        if "leaderboard" in kpis:
            lb_df = kpis["leaderboard"].reset_index()
            fig = px.bar(
                lb_df.sort_values("Revenue_Won", ascending=True),
                x="Revenue_Won",
                y="Sales_Rep",
                orientation="h",
                hover_data=["Total_Leads", "Won_Deals", "Win_Rate", "Avg_Deal_Size"],
                title="Revenue Won by Sales Rep",
                labels={"Revenue_Won": "Revenue Won (£)", "Sales_Rep": "Sales Rep"},
            )
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(
                lb_df.style.format({
                    "Revenue_Won": "£{:,.0f}",
                    "Avg_Deal_Size": "£{:,.0f}",
                    "Win_Rate": "{:.1%}",
                }),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("Sales Rep column not found in this file.")

    with tab2:
        if "by_lead_source" in kpis:
            source_df = kpis["by_lead_source"].reset_index()
            fig = px.bar(
                source_df,
                x="Lead_Source",
                y="Revenue_Won",
                hover_data=["Total_Leads", "Won", "Win_Rate"],
                title="Revenue Won by Lead Source",
                labels={"Revenue_Won": "Revenue Won (£)", "Lead_Source": "Lead Source"},
            )
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(
                source_df.style.format({
                    "Revenue_Won": "£{:,.0f}",
                    "Win_Rate": "{:.1%}",
                }),
                use_container_width=True,
            )
        else:
            st.info("Lead Source column not found in this file.")

    with tab3:
        if "by_meeting_class" in kpis:
            class_df = kpis["by_meeting_class"].reset_index()
            fig = px.bar(
                class_df,
                x="Meeting_Class",
                y="Revenue_Won",
                hover_data=["Total_Leads", "Won_Deals", "Win_Rate"],
                title="Revenue Won by Event Type",
                labels={"Revenue_Won": "Revenue Won (£)", "Meeting_Class": "Event Type"},
            )
            st.plotly_chart(fig, use_container_width=True)
            st.caption(
                "Note: high revenue doesn't always mean high win rate — an event type "
                "can attract lots of enquiries but convert poorly, which the win rate column reveals."
            )
            st.dataframe(
                class_df.style.format({
                    "Revenue_Won": "£{:,.0f}",
                    "Win_Rate": "{:.1%}",
                }),
                use_container_width=True,
            )
        else:
            st.info("Meeting Class column not found in this file.")

    with tab4:
        if kpis["pipeline_aging"] is not None:
            aging_df = kpis["pipeline_aging"].reset_index()
            fig = px.bar(
                aging_df,
                x="Aging_Bucket",
                y="Open_Value",
                hover_data=["Open_Deals"],
                title="Open Pipeline Value by Age (Tentative + Prospect bookings)",
                labels={"Open_Value": "Open Pipeline Value (£)", "Aging_Bucket": "Days Since Enquiry"},
            )
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(
                aging_df.style.format({"Open_Value": "£{:,.0f}"}),
                use_container_width=True,
            )
        else:
            st.info("No open (Tentative/Prospect) bookings in this selection.")

    with tab5:
        if kpis.get("seasonality") is not None:
            season_df = kpis["seasonality"].reset_index()
            fig = px.bar(
                season_df,
                x="Month_Name",
                y="Revenue_Won",
                hover_data=["Total_Bookings", "Won_Deals"],
                title="Revenue Won by Event Month (ordered by Financial Year, Jul–Jun)",
                labels={"Revenue_Won": "Revenue Won (£)", "Month_Name": "Event Month"},
            )
            fig.update_xaxes(categoryorder="array", categoryarray=season_df["Month_Name"])
            st.plotly_chart(fig, use_container_width=True)
            st.caption(
                "Based on the Arrival (event) date. Useful for spotting peak vs. "
                "off-peak months for staffing and pricing decisions."
            )
            st.dataframe(
                season_df.style.format({"Revenue_Won": "£{:,.0f}"}),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("Arrival Date column not found in this file.")

    with tab6:
        if kpis.get("revenue_trend") is not None:
            trend_df = kpis["revenue_trend"].reset_index()
            fig = px.line(
                trend_df,
                x="Year_Month",
                y="Revenue_Won",
                markers=True,
                title="Revenue Won by Month (Event Date)",
                labels={"Revenue_Won": "Revenue Won (£)", "Year_Month": "Month"},
            )
            st.plotly_chart(fig, use_container_width=True)
            st.caption(
                "Shows revenue momentum over time — including future months where "
                "events are already booked as Definite. A single financial-year "
                "snapshot can hide a trend like this."
            )
            st.dataframe(
                trend_df.style.format({"Revenue_Won": "£{:,.0f}"}),
                use_container_width=True,
            )
        else:
            st.info("Arrival Date column not found in this file.")

    with tab7:
        if kpis.get("lost_reasons") is not None:
            reasons_df = kpis["lost_reasons"].reset_index()
            fig = px.bar(
                reasons_df.head(15),
                x="Count",
                y="Lost_Reason",
                orientation="h",
                hover_data=["Value"],
                title="Top Reasons for Lost / Turned Down Bookings",
                labels={"Count": "Number of Bookings", "Lost_Reason": "Reason"},
            )
            fig.update_yaxes(categoryorder="total ascending")
            st.plotly_chart(fig, use_container_width=True)
            st.caption(
                "Reasons prefixed 'LT-' are Lost bookings (client chose someone else); "
                "'TD-' are Turned Down bookings (The Brewery declined to host). "
                "This is probably the most actionable breakdown here — it shows *why* "
                "you're losing business, not just how much."
            )
            st.dataframe(
                reasons_df.style.format({"Value": "£{:,.0f}"}),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("Lost Reason column not found in this file, or no Lost/TurnedDown bookings in this selection.")

    with tab8:
        if kpis.get("top_accounts") is not None:
            acct_df = kpis["top_accounts"]
            fig = px.bar(
                acct_df.sort_values("Revenue_Won", ascending=True),
                x="Revenue_Won",
                y="Account",
                orientation="h",
                title="Top 10 Accounts by Won Revenue",
                labels={"Revenue_Won": "Revenue Won (£)", "Account": "Account"},
            )
            st.plotly_chart(fig, use_container_width=True)
            st.metric(
                "Top 10 Accounts' Share of Total Won Revenue",
                pct(kpis["top10_account_share"]),
                help="A high share means more revenue risk if one of these accounts leaves.",
            )
            st.dataframe(
                acct_df.style.format({"Revenue_Won": "£{:,.0f}"}),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("Account column not found in this file.")

    st.divider()

    # ---------------- Raw data preview ----------------
    with st.expander("View raw data"):
        st.dataframe(df, use_container_width=True)


if __name__ == "__main__":
    main()

